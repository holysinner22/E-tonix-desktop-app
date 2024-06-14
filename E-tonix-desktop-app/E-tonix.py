import tkinter as tk
from tkinter import messagebox, ttk
import sqlite3
from datetime import datetime
from fpdf import FPDF
from tkinter import filedialog
import openpyxl

class LoginWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("Login")
        self.geometry("1600x1270")
        self.configure(bg='#2c3e50')  # Set background color

        # Create a style for the labels and buttons
        style = ttk.Style()
        style.configure('TLabel', background='#2c3e50', foreground='#ecf0f1', font=('Georgia', 12, 'bold'))
        style.configure('TEntry', font=('Helvetica', 12))
        style.configure('TButton', background='#3498db', foreground='#E89110', font=('Georgia', 12, 'bold'))

        self.welcome_label = ttk.Label(self, text="Welcome to E-Tonix", font=('Georgia', 16, 'bold'), background='#2c3e50', foreground='#ecf0f1')
        self.welcome_label.pack(pady=30)

        self.username_label = ttk.Label(self, text="Username")
        self.username_label.pack(pady=5)
        self.username_entry = ttk.Entry(self)
        self.username_entry.pack(pady=5)

        self.password_label = ttk.Label(self, text="Password")
        self.password_label.pack(pady=5)
        self.password_entry = ttk.Entry(self, show="*")
        self.password_entry.pack(pady=5)

        self.login_button = ttk.Button(self, text="Login", command=self.login)
        self.login_button.pack(pady=5)

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        if username == "admin" and password == "AdminE":
            self.parent.show_main_app()
            self.destroy()
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")

class DesktopApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.withdraw()  # Hide main window until login is successful
        self.title("E-Tonix")
        self.login_window = LoginWindow(self)

        self.home_frame = ttk.Frame(self)
        self.inventory_frame = ttk.Frame(self)
        self.pending_tasks_frame = ttk.Frame(self)

        self.create_home_table()
        self.create_inventory_table()
        self.create_pending_tasks_table()
        self.create_home_buttons()
        self.create_inventory_buttons()
        self.create_search_fields()

        self.create_database()
        self.load_home_data()
        self.create_inventory_database()
        self.load_inventory_data()
        self.load_pending_tasks_data()

        self.menu = tk.Menu(self)
        self.menu.add_command(label="Home", command=self.show_home)
        self.menu.add_command(label="Inventory", command=self.show_inventory)
        self.menu.add_command(label="Pending Tasks", command=self.show_pending_tasks)
        self.config(menu=self.menu)

        self.update_date()

    def show_main_app(self):
        self.deiconify()  # Show main window after successful login
        self.show_home()  # Show home by default after login.

    def generate_daily_report(self):
        current_date = datetime.now().strftime("%Y-%m-%d")
        excel_filename = f"daily_report_{current_date}.xlsx"
        pdf_filename = f"daily_report_{current_date}.pdf"

        # Export data to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        # Add headers
        headers = ["Arrival Date", "Name", "Phone", "Device", "Issues", "Description", "Status", 
                   "Technician", "Solution", "Checkup Price", "Price Charged", "Payment Method", "Due Date", "Total Income"]
        ws.append(headers)

        # Fetch data for the current date from the database
        conn = sqlite3.connect("etonic.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM home_data WHERE arrival_date=?", (current_date,))
        rows = cursor.fetchall()

        for row in rows:
            row = row[1:]  # Omit the ID column by skipping the first element
            checkup_price = float(row[9].replace(',', '')) if row[9] else 0
            price_charged = float(row[10].replace(',', '')) if row[10] else 0
            total_income = checkup_price + price_charged
            row_data = list(row) + [total_income]  # Append total income to the row
            ws.append(row_data)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        wb.save(excel_filename)

        # Convert Excel to PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Read data from Excel and write to PDF
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                pdf.cell(40, 10, str(cell), border=1)
            pdf.ln()

        # Output PDF to file
        pdf.output(pdf_filename)

        messagebox.showinfo("Success", f"Daily report saved to {pdf_filename}")

        # Close the database connection
        conn.close()

    def get_current_date(self):
        return datetime.now().strftime("%Y-%m-%d -- %H:%M:%S")

    def update_date(self):
        current_date = self.get_current_date()
        self.date_label_home.config(text=f"Current Date and Time: {current_date}")
        self.date_label_inventory.config(text=f"Current Date and Time: {current_date}")
        self.date_label_pending_tasks.config(text=f"Current Date and Time: {current_date}")
        self.after(1000, self.update_date)

    def create_home_table(self):
        self.table_frame = ttk.Frame(self.home_frame)
        self.table_frame.pack(fill=tk.BOTH, expand=True)

        self.date_label_home = ttk.Label(self.home_frame, text=f"Current Date and Time: {self.get_current_date()}")
        self.date_label_home.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.table = ttk.Treeview(self.table_frame, columns=("ID","Arrival Date", "Name", "Phone", "Device", "Issues","Description", "Status", "Technician", "Solution", "Checkup Price", "Price Charged", "Payment Method", "Due Date"))
        self.table.heading("#0", text="")
        self.table.heading("ID", text="ID")
        self.table.heading("Arrival Date", text="Arrival Date")
        self.table.heading("Name", text="Name")
        self.table.heading("Phone", text="Phone")
        self.table.heading("Device", text="Device")
        self.table.heading("Issues", text="Issues")
        self.table.heading("Description", text="Description")
        self.table.heading("Status", text="Status")
        self.table.heading("Technician", text="Technician")
        self.table.heading("Solution", text="Solution")
        self.table.heading("Checkup Price", text="Checkup Price")
        self.table.heading("Price Charged", text="Price Charged")
        self.table.heading("Payment Method", text="Payment Method")
        self.table.heading("Due Date", text="Due Date")

        self.table_scroll_y = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.table.yview)
        self.table_scroll_x = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.table.xview)

        self.table.configure(yscrollcommand=self.table_scroll_y.set, xscrollcommand=self.table_scroll_x.set)

        self.table_scroll_y.pack(side="right", fill="y")
        self.table_scroll_x.pack(side="bottom", fill="x")

        self.table.pack(expand=True, fill=tk.BOTH)

    def create_inventory_table(self):
        self.inventory_table_frame = ttk.Frame(self.inventory_frame)
        self.inventory_table_frame.pack(fill=tk.BOTH, expand=True)

        self.date_label_inventory = ttk.Label(self.inventory_frame, text=f"Current Date and Time: {self.get_current_date()}")
        self.date_label_inventory.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.inventory_table = ttk.Treeview(self.inventory_table_frame, columns=("ID", "Hardware Available", "Total", "Status"))
        self.inventory_table.heading("#0", text="")
        self.inventory_table.heading("ID", text="ID")
        self.inventory_table.heading("Hardware Available", text="Hardware Available")
        self.inventory_table.heading("Total", text="Total")
        self.inventory_table.heading("Status", text="Status")

        self.inventory_scroll_y = ttk.Scrollbar(self.inventory_table_frame, orient="vertical", command=self.inventory_table.yview)
        self.inventory_scroll_x = ttk.Scrollbar(self.inventory_table_frame, orient="horizontal", command=self.inventory_table.xview)

        self.inventory_table.configure(yscrollcommand=self.inventory_scroll_y.set, xscrollcommand=self.inventory_scroll_x.set)

        self.inventory_scroll_y.pack(side="right", fill="y")
        self.inventory_scroll_x.pack(side="bottom", fill="x")

        self.inventory_table.pack(expand=True, fill=tk.BOTH)

    def create_pending_tasks_table(self):
        self.pending_tasks_table_frame = ttk.Frame(self.pending_tasks_frame)
        self.pending_tasks_table_frame.pack(fill=tk.BOTH, expand=True)

        self.date_label_pending_tasks = ttk.Label(self.pending_tasks_frame, text=f"Current Date and Time: {self.get_current_date()}")
        self.date_label_pending_tasks.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.pending_tasks_table = ttk.Treeview(self.pending_tasks_table_frame, columns=("ID","Arrival Date", "Name", "Phone", "Device", "Issues","Description", "Status", "Technician", "Solution", "Checkup Price", "Price Charged", "Payment Method", "Due Date"))
        self.pending_tasks_table.heading("#0", text="")
        self.pending_tasks_table.heading("ID", text="ID")
        self.pending_tasks_table.heading("Arrival Date", text="Arrival Date")
        self.pending_tasks_table.heading("Name", text="Name")
        self.pending_tasks_table.heading("Phone", text="Phone")
        self.pending_tasks_table.heading("Device", text="Device")
        self.pending_tasks_table.heading("Issues", text="Issues")
        self.pending_tasks_table.heading("Description", text="Description")
        self.pending_tasks_table.heading("Status", text="Status")
        self.pending_tasks_table.heading("Technician", text="Technician")
        self.pending_tasks_table.heading("Solution", text="Solution")
        self.pending_tasks_table.heading("Checkup Price", text="Checkup Price")
        self.pending_tasks_table.heading("Price Charged", text="Price Charged")
        self.pending_tasks_table.heading("Payment Method", text="Payment Method")
        self.pending_tasks_table.heading("Due Date", text="Due Date")

        self.pending_tasks_table_scroll_y = ttk.Scrollbar(self.pending_tasks_table_frame, orient="vertical", command=self.pending_tasks_table.yview)
        self.pending_tasks_table_scroll_x = ttk.Scrollbar(self.pending_tasks_table_frame, orient="horizontal", command=self.pending_tasks_table.xview)

        self.pending_tasks_table.configure(yscrollcommand=self.pending_tasks_table_scroll_y.set, xscrollcommand=self.pending_tasks_table_scroll_x.set)

        self.pending_tasks_table_scroll_y.pack(side="right", fill="y")
        self.pending_tasks_table_scroll_x.pack(side="bottom", fill="x")

        self.pending_tasks_table.pack(expand=True, fill=tk.BOTH)

    def create_home_buttons(self):
        self.button_frame = ttk.Frame(self.home_frame)
        self.button_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.new_button = ttk.Button(self.button_frame, text="New", command=self.open_new_entry_form)

        self.new_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.edit_button = ttk.Button(self.button_frame, text="Edit", command=self.edit_entry)
        self.edit_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.delete_button = ttk.Button(self.button_frame, text="Delete", command=self.delete_entry)
        self.delete_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.download_button = ttk.Button(self.button_frame, text="Download", command=self.download_entry)
        self.download_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.daily_report_button = ttk.Button(self.button_frame, text="Generate Daily Report", command=self.generate_daily_report)
        self.daily_report_button.pack(side=tk.LEFT, padx=5, pady=5)

    def create_inventory_buttons(self):
        self.inventory_button_frame = ttk.Frame(self.inventory_frame)
        self.inventory_button_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.inventory_new_button = ttk.Button(self.inventory_button_frame, text="New", command=self.open_new_inventory_entry_form)
        self.inventory_new_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.inventory_edit_button = ttk.Button(self.inventory_button_frame, text="Edit", command=self.edit_inventory_entry)
        self.inventory_edit_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.inventory_delete_button = ttk.Button(self.inventory_button_frame, text="Delete", command=self.delete_inventory_entry)
        self.inventory_delete_button.pack(side=tk.LEFT, padx=5, pady=5)

    def create_search_fields(self):
        self.search_frame_home = ttk.Frame(self.home_frame)
        self.search_frame_home.pack(side=tk.TOP, fill=tk.X)
        self.search_entry_home = ttk.Entry(self.search_frame_home)
        self.search_entry_home.pack(side=tk.LEFT, padx=5, pady=5)
        self.search_button_home = ttk.Button(self.search_frame_home, text="Search", command=self.search_home)
        self.search_button_home.pack(side=tk.LEFT, padx=5, pady=5)
        self.clear_button_home = ttk.Button(self.search_frame_home, text="Clear Search Results", command=self.reset_table)
        self.clear_button_home.pack(side=tk.LEFT, padx=5, pady=5)

        self.search_frame_inventory = ttk.Frame(self.inventory_frame)
        self.search_frame_inventory.pack(side=tk.TOP, fill=tk.X)
        self.search_entry_inventory = ttk.Entry(self.search_frame_inventory)
        self.search_entry_inventory.pack(side=tk.LEFT, padx=5, pady=5)
        self.search_button_inventory = ttk.Button(self.search_frame_inventory, text="Search", command=self.search_inventory)
        self.search_button_inventory.pack(side=tk.LEFT, padx=5, pady=5)
        self.clear_button_inventory = ttk.Button(self.search_frame_inventory, text="Clear Search Results", command=self.reset_table)
        self.clear_button_inventory.pack(side=tk.LEFT, padx=5, pady=5)

    def open_new_entry_form(self):
        NewEntryForm(self)

    def open_new_inventory_entry_form(self):
        NewEntryForm(self, is_inventory=True)

    def edit_entry(self):
        selected_item = self.table.selection()
        if selected_item:
            item = self.table.item(selected_item)
            data = item["values"]
            NewEntryForm(self, data=data)

    def delete_entry(self):
        selected_item = self.table.selection()
        if selected_item:
            item = self.table.item(selected_item)
            item_id = item["values"][0]
            self.cursor.execute("DELETE FROM home_data WHERE id=?", (item_id,))
            self.conn.commit()
            self.table.delete(selected_item)

    def download_entry(self):
        selected_item = self.table.selection()
        if selected_item:
            item = self.table.item(selected_item)
            item_id = item["values"][0]

            try:
                # Fetch the row data from the database
                self.cursor.execute("SELECT * FROM home_data WHERE id=?", (item_id,))
                row_data = self.cursor.fetchone()

                if row_data:
                    # Assuming you know the order of the columns
                    column_names = [desc[0] for desc in self.cursor.description]

                    # Create a dictionary for row data
                    row_dict = dict(zip(column_names, row_data))

                    # Save the data to a PDF file
                    save_path = f"row_{row_dict['id']}.pdf"

                    # Create a PDF object
                    pdf = FPDF()
                    pdf.add_page()
                    pdf.set_font("Arial", size=12)

                    # Add row data to the PDF
                    for key, value in row_dict.items():
                        pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)

                    # Output the PDF to a file
                    pdf.output(save_path)

                    messagebox.showinfo("Success", f"Row data saved to {save_path}")
                else:
                    messagebox.showerror("Error", "No data found for the selected item.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save row data: {e}")
        else:
            messagebox.showwarning("Warning", "No item selected.")

    def edit_inventory_entry(self):
        selected_item = self.inventory_table.selection()
        if selected_item:
            item = self.inventory_table.item(selected_item)
            data = item["values"]
            NewEntryForm(self, data=data, is_inventory=True)

    def delete_inventory_entry(self):
        selected_item = self.inventory_table.selection()
        if selected_item:
            item = self.inventory_table.item(selected_item)
            item_id = item["values"][0]
            self.inventory_cursor.execute("DELETE FROM inventory_data WHERE id=?", (item_id,))
            self.inventory_conn.commit()
            self.inventory_table.delete(selected_item)

    def reset_table(self):
        self.load_home_data()
        self.load_inventory_data()
        self.load_pending_tasks_data()
        self.search_entry_home.delete(0, tk.END)
        self.search_entry_inventory.delete(0, tk.END)

    def search_home(self):
        query = self.search_entry_home.get().strip()
        if query:
            self.cursor.execute("""
                SELECT * FROM home_data
                WHERE arrival_date LIKE ? OR name LIKE ? OR phone LIKE ? OR device LIKE ? OR issues LIKE ? OR description LIKE ? OR status LIKE ? OR technician LIKE ? OR solution LIKE ? OR checkup_price LIKE ? OR price_charged LIKE ? OR payment_method LIKE ? OR due_date LIKE ? 
            """, (f'%{query}%',) * 13)
            rows = self.cursor.fetchall()
            self.table.delete(*self.table.get_children())
            for row in rows:
                self.table.insert("", "end", values=row)

    def search_inventory(self):
        query = self.search_entry_inventory.get().strip()
        if query:
            self.inventory_cursor.execute("""
                SELECT * FROM inventory_data
                WHERE hardware_available LIKE ? OR total LIKE ? OR status LIKE ?
            """, (f'%{query}%',) * 3)
            rows = self.inventory_cursor.fetchall()
            self.inventory_table.delete(*self.inventory_table.get_children())
            for row in rows:
                self.inventory_table.insert("", "end", values=row)

    def search_pending_tasks(self):
        query = self.search_entry_pending_tasks_data.get().strip()
        if query:
            self.pending_tasks_cursor.execute("""
                SELECT * FROM pending_tasks_data
                WHERE arrival_date LIKE ? OR name LIKE ? OR phone LIKE ? OR device LIKE ? OR issues LIKE ? OR description LIKE ? OR status LIKE ? OR technician LIKE ? OR solution LIKE ? OR checkup_price LIKE ? OR price_charged LIKE ? OR payment_method LIKE ? OR due_date LIKE ? 
            """, (f'%{query}%',) * 13)
            rows = self.pending_tasks_cursor.fetchall()
            self.pending_tasks_table.delete(*self.pending_tasks_table.get_children())
            for row in rows:
                self.pending_tasks_table.insert("", "end", values=row)

    def show_home(self):
        self.home_frame.pack(fill=tk.BOTH, expand=True)
        self.inventory_frame.pack_forget()
        self.pending_tasks_frame.pack_forget()
        self.reset_table()

    def show_inventory(self):
        self.inventory_frame.pack(fill=tk.BOTH, expand=True)
        self.home_frame.pack_forget()
        self.pending_tasks_frame.pack_forget()
        self.reset_table()

    def show_pending_tasks(self):
        self.pending_tasks_frame.pack(fill=tk.BOTH, expand=True)
        self.home_frame.pack_forget()
        self.inventory_frame.pack_forget()
        self.reset_table()

    def create_database(self):
        self.conn = sqlite3.connect("etonic.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS home_data (
                id INTEGER PRIMARY KEY,
                arrival_date TEXT,
                name TEXT,
                phone TEXT,
                device TEXT,
                issues TEXT,
                description TEXT,
                status TEXT,
                technician TEXT,
                solution TEXT,
                checkup_price TEXT,
                price_charged TEXT,
                payment_method TEXT,
                due_date TEXT
            )
        """)
        self.conn.commit()

    def load_home_data(self):
        self.table.delete(*self.table.get_children())
        self.cursor.execute("SELECT * FROM home_data")
        rows = self.cursor.fetchall()
        for row in rows:
            self.table.insert("", "end", values=row)

    def insert_home_data(self, data):
        self.cursor.execute("""
            INSERT INTO home_data (arrival_date, name, phone, device, issues, description, status, technician, solution, checkup_price, price_charged, payment_method, due_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, data)
        self.conn.commit()
        self.load_home_data()

    def update_home_data(self, data):
        self.cursor.execute("""
            UPDATE home_data
            SET arrival_date=?, name=?, phone=?, device=?, issues=?, description=?, status=?, technician=?, solution=?, checkup_price=?, price_charged=?, payment_method=?, due_date=?
            WHERE id=?
        """, data[1:] + [data[0]])
        self.conn.commit()
        self.load_home_data()

    def create_inventory_database(self):
        self.inventory_conn = sqlite3.connect("inventory.db")
        self.inventory_cursor = self.inventory_conn.cursor()
        self.inventory_cursor.execute("""
            CREATE TABLE IF NOT EXISTS inventory_data (
                id INTEGER PRIMARY KEY,
                hardware_available TEXT,
                total INTEGER,
                status TEXT
            )
        """)
        self.inventory_conn.commit()

    def load_inventory_data(self):
        self.inventory_table.delete(*self.inventory_table.get_children())
        self.inventory_cursor.execute("SELECT * FROM inventory_data")
        rows = self.inventory_cursor.fetchall()
        for row in rows:
            self.inventory_table.insert("", "end", values=row)

    def insert_inventory_data(self, data):
        self.inventory_cursor.execute("""
            INSERT INTO inventory_data (hardware_available, total, status)
            VALUES (?, ?, ?)
        """, data)
        self.inventory_conn.commit()
        self.load_inventory_data()

    def update_inventory_data(self, data):
        self.inventory_cursor.execute("""
            UPDATE inventory_data
            SET hardware_available=?, total=?, status=?
            WHERE id=?
        """, data[1:] + [data[0]])
        self.inventory_conn.commit()
        self.load_inventory_data()

    def load_pending_tasks_data(self):
        conn = sqlite3.connect("etonic.db")
        cursor = conn.cursor()
        self.pending_tasks_table.delete(*self.pending_tasks_table.get_children())
        cursor.execute("SELECT * FROM home_data WHERE status ='pending'")
        rows = cursor.fetchall()
        for row in rows:
            self.pending_tasks_table.insert("", "end", values=row)
        conn.close()


class NewEntryForm(tk.Toplevel):
    def __init__(self, parent, data=None, is_inventory=False):
        super().__init__(parent)
        self.parent = parent
        self.is_inventory = is_inventory
        self.data = data
        self.title("New Entry" if data is None else "Edit Entry")

        if is_inventory:
            self.entry_labels = ["Hardware Available", "Total", "Status"]
        else:
            self.entry_labels = [
                "Arrival Date", "Name", "Phone", "Device", "Issues", "Description", "Status", "Technician", 
                "Solution", "Checkup Price", "Price Charged", "Payment Method", "Due Date"
            ]

        self.entry_fields = {}
        row = 0
        for label_text in self.entry_labels:
            label = ttk.Label(self, text=label_text)
            label.grid(row=row, column=0, padx=5, pady=5, sticky="w")
            entry = ttk.Entry(self)
            if label_text == "Arrival Date" and not data:  # Set default value for Arrival Date if adding a new entry
                entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
            entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            self.entry_fields[label_text] = entry
            
            if label_text == "Phone" and not data:  # Set default value for Phone if adding a new entry
                entry.insert(0, "+255")
            entry.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            self.entry_fields[label_text] = entry
            row += 1
            
        if data:
            for label_text, value in zip(self.entry_labels, data[1:]):  # Skip the ID
                self.entry_fields[label_text].insert(0, value)

        if is_inventory:
            self.save_button = ttk.Button(self, text="Save", command=self.save_inventory_data)
        else:
            self.save_button = ttk.Button(self, text="Save", command=self.save_home_data)
        self.save_button.grid(row=row, columnspan=2, padx=5, pady=5)

    def save_home_data(self):
        data = [self.entry_fields[label].get() for label in self.entry_labels]
        if self.data:  # if editing
            data = [self.data[0]] + data  # preserve ID
            self.parent.update_home_data(data)
        else:
            self.parent.insert_home_data(data)
        self.destroy()

    def save_inventory_data(self):
        data = [self.entry_fields[label].get() for label in self.entry_labels]
        if self.data:  # if editing
            data = [self.data[0]] + data  # preserve ID
            self.parent.update_inventory_data(data)
        else:
            try:
                data[1] = int(data[1])  # Convert total to integer
            except ValueError:
                messagebox.showerror("Invalid Input", "Total must be an integer")
                return
            self.parent.insert_inventory_data(data)
        self.destroy()

if __name__ == "__main__":
    app = DesktopApp()
    app.mainloop()

